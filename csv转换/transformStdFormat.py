# encoding=utf-8
import os
import openpyxl
import sys
from openpyxl.cell.cell import MergedCell
import pandas as pd
import datetime





class transformStdFormat():
    def __init__(self,file_path,file_name=None):
        self.file_path = file_path
        # 获取文件后缀格式
        self.file_object = file_path.split('.')[-1]
        # 判断传入文件是否支持转换
        self.file_is_suppost()
        # 获取文件名称
        if file_name == None:
            self.file_name = os.path.basename(self.file_path)
        else:
            self.basename = file_name
        # 获取文件父级目录
        self.parent_path = os.path.dirname(self.file_path)
        print(self.parent_path)

    # 判断文件类型是否支持本脚本的转换
    def file_is_suppost(self):
        if self.file_object not in ['csv','xlsx','xls']:
            print(f'当前转换脚本不支持文件后缀为{self.file_object}的文件，程序主动终止！')
            sys.exit()

    # 处理转换文件
    def run(self,onlySelect):
        '''
            onlySelect : 是否只是查看
        '''
        if onlySelect not in [True,False,1,0]:
            raise ValueError('请输入onlySelect参数正确的取值值：【True,False】or【1,0】')

        excell = openpyxl.load_workbook(self.file_path)
        self.allSheetNames = excell.sheetnames
        print('当前处理表格中所有的sheet页：', self.allSheetNames)
        print('\n')

        # 循环每一个sheet页的名字
        for sheet_name in excell.sheetnames:
            self.currentSheetNames = sheet_name
            print(f'----当前处理sheet页为：{sheet_name}')

            # 将sheet页转换成excell对象
            table = excell[sheet_name]
            # 每一行对象
            sheet = tuple(table.rows)

            all_row = []
            for c, row in enumerate(sheet):
                row_data = []
                for cell in row:
                    if type(cell) == MergedCell:
                        break
                    else:
                        row_data.append(cell.value)
                # 过滤行为单一值的数据
                if len(row_data) != 1:
                    all_row.append(row_data)

            self.sheet_df = pd.DataFrame(all_row)
            # sheet_df.dropna(inplace=True)
            self.sheet_df.drop_duplicates(keep='first', inplace=True)
            self.sheet_df.reset_index(drop=True, inplace=True)
            self.sheet_df.columns = self.sheet_df.iloc[0, :].tolist()
            self.sheet_df.drop(labels=0, inplace=True)

            # 判断返回形式
            if onlySelect == True:
                print(f'{self.currentSheetNames}抽取数据如下所示：\n\n',self.sheet_df)
            else:
                # 输出文件
                if self.file_object in ['xlsx','xls']:
                    self.sheet_df.to_excel(rf'{self.parent_path}\result_{sheet_name}.{self.file_object}', index=False, encoding='utf-8')
                elif self.file_object == 'csv':
                    self.sheet_df.to_csv(rf'{self.parent_path}\result_{sheet_name}.{self.file_object}', index=False, encoding='utf-8')
                print('sheet文件已输出')




if __name__ == '__main__':

    file_path = r'C:\Users\dell\Desktop\江苏高科技投资集团有限公司.xlsx'


    transform = transformStdFormat(file_path)
    transform.run(onlySelect=False)

