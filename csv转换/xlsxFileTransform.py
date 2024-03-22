import os
import openpyxl
import sys
from openpyxl.cell.cell import MergedCell
import pandas as pd
import datetime



def process(file_path):
    # 循环文件夹中的csv和xlsx文件
    # for fileName in os.listdir(file_path):
        # if fileName.split('.')[-1] in ['csv','xlsx']:
        #     print(f'++++++++++++++++++++++当前处理 {fileName} 文件++++++++++++++++++++++')
    if 1 ==1:
        if 1 == 1:
            # fileName = r'D:\Gitlab\my_world\人行csv转换\附件1-3-20221115.xlsx'
            fileName = r'C:\Users\dell\Desktop\江苏高科技投资集团有限公司.xlsx'
            excell = openpyxl.load_workbook(fileName)
            print('当前处理表格中sheet页：',excell.sheetnames)

            # 循环每一个sheet页的名字
            for sheet_name in excell.sheetnames:
                print(f'----处理：{sheet_name}')
                table = excell[sheet_name]
                sheet = tuple(table.rows)

                all_row = []
                for c,row in enumerate(sheet):
                    row_data = []
                    for cell in row:
                        if type(cell) == MergedCell:
                            break
                        else:
                            row_data.append(cell.value)
                    # 过滤行为单一值的数据
                    if len(row_data) != 1:
                        all_row.append(row_data)

                sheet_df = pd.DataFrame(all_row)
                # sheet_df.dropna(inplace=True)
                sheet_df.drop_duplicates(keep='first',inplace=True)
                sheet_df.reset_index(drop=True,inplace=True)
                sheet_df.columns = sheet_df.iloc[0,:].tolist()
                sheet_df.drop(labels=0,inplace=True)

                print(sheet_df)
                sheet_df.to_csv(f'result_{sheet_name}.csv',index=False,encoding='utf-8')



if __name__ == '__main__':

    csv_path = r'/csv转换'

    process(csv_path)




















