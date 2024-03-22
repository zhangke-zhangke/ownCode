import os
import openpyxl
import sys
from openpyxl.cell.cell import MergedCell
import pandas as pd


# 解决会以科学计数法写入csv
def num_out(data):
    data = str(data) + '\t'
    return data


def process(file_path):
    # 循环文件夹中的xlsx文件
    for fileName in os.listdir(file_path):
        if fileName.split('.')[-1] == 'xlsx':
            print(f'++++++++++++++++++++++当前处理 {fileName} 文件++++++++++++++++++++++')
            excell = openpyxl.load_workbook(fileName)
            print(excell.sheetnames)

            # 循环每一个sheet页的名字
            for sheet_name in excell.sheetnames:
                table = excell[sheet_name]
                sheet = tuple(table.rows)

                all_row = []
                for c,row in enumerate(sheet):
                    row_data = []
                    for cell in row:
                        if type(cell) == MergedCell:
                            break
                        else:
                            row_data.append(cell.value)
                    # 过滤行为单一值的数据
                    if len(row_data) != 1:
                        all_row.append(row_data)
                # 封装为pandas数据对象
                data = pd.DataFrame(all_row).dropna()
                # print(data.columns)
                data = data.drop_duplicates(keep='first').reset_index(drop=True)
                data.columns = data.iloc[0,:].tolist()
                data.drop(labels=0,inplace=True)

                # 按原数据写入，避免科学记数法写入
                for i in data.columns:
                    data[i] = data[i].map(num_out)
                print(data)
                # 写入csv
                data.to_csv(f'result_{sheet_name}.csv',index=False,encoding='utf-8')





if __name__ == '__main__':

    csv_path = r'D:\Gitlab\my_world\人行csv转换'

    process(csv_path)




















